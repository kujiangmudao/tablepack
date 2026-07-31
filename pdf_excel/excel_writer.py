# -*- coding: utf-8 -*-
"""Write multi-sheet workbooks from TableItem list."""

from __future__ import annotations

import re

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .clean import clean_cell_text
from .html_table import html_table_to_grid
from .models import TableItem


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    # Excel forbids: : \ / ? * [ ] and names longer than 31; also avoid leading/trailing '
    name = re.sub(r"[:\\/?*\[\]]", "_", name or "")
    name = name.replace("\x00", "")
    name = name.strip(" '\"\t\r\n") or "Table"
    if len(name) > 31:
        name = name[:31].rstrip(" '\"")
    if not name:
        name = "Table"
    # Excel reserves the name "History" in some contexts — avoid collision quietly
    if name.lower() == "history":
        name = "History_"
    base = name
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    used.add(name)
    return name


def write_excel(tables: list[TableItem], xlsx_path) -> list[str]:
    """Write all tables into one workbook (one sheet per table)."""
    from pathlib import Path

    xlsx_path = Path(xlsx_path)
    issues: list[str] = []
    wb = Workbook()
    default = wb.active
    used_names: set[str] = set()
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

    if not tables:
        default.title = "无表格"
        default["A1"] = "未从该 PDF 中识别到可转换的表格"
        issues.append("no tables found")
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)
        return issues

    first = True
    for t in tables:
        grid, g_issues = html_table_to_grid(t.html_body)
        t.issues.extend(g_issues)
        if g_issues:
            issues.append(f"表{t.index} ({t.caption or '无标题'}): {'; '.join(g_issues)}")

        cap = t.caption or f"第{t.page_idx + 1}页表格"
        cap_short = re.split(r"\s*Table\s+\d+", cap, maxsplit=1, flags=re.I)[0].strip()
        if not cap_short:
            cap_short = cap
        sheet_name = sanitize_sheet_name(f"表{t.index}_{cap_short}", used_names)

        if first:
            ws = default
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = f"标题: {t.caption or '(无)'}"
        ws["A2"] = f"页码: {t.page_idx + 1}"
        if t.footnote:
            ws["A3"] = "注释: " + " | ".join(
                clean_cell_text(BeautifulSoup(str(f), "lxml").get_text(" ", strip=True)) for f in t.footnote
            )
        start_row = 5 if t.footnote else 4

        if not grid:
            ws.cell(row=start_row, column=1, value="(表格 HTML 解析失败，请查看原始表格图片与问题说明)")
            continue

        for r_i, row in enumerate(grid):
            for c_i, val in enumerate(row):
                cell = ws.cell(row=start_row + r_i, column=c_i + 1, value=val)
                cell.border = thin
                cell.alignment = wrap
                if r_i == 0:
                    cell.font = header_font

        for c_i in range(len(grid[0]) if grid else 1):
            max_len = 8
            for row in grid:
                if c_i < len(row) and row[c_i]:
                    s = str(row[c_i])
                    max_len = max(max_len, min(40, int(len(s) * 1.2) + 2))
            ws.column_dimensions[get_column_letter(c_i + 1)].width = max_len

        ws.sheet_properties.tabColor = "4472C4"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return issues
