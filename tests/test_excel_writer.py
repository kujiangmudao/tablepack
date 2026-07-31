# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import load_workbook

from pdf_excel.excel_writer import sanitize_sheet_name, write_excel
from pdf_excel.models import TableItem


def test_sanitize_sheet_name_unique():
    used: set[str] = set()
    a = sanitize_sheet_name("表1_测试:斜杠/星*", used)
    b = sanitize_sheet_name("表1_测试:斜杠/星*", used)
    assert a != b
    assert len(a) <= 31
    assert ":" not in a and "/" not in a


def test_write_excel_simple(tmp_path: Path):
    tables = [
        TableItem(
            index=1,
            page_idx=0,
            caption="表1 演示",
            caption_raw=["表1 演示"],
            footnote=[],
            html_body="<table><tr><th>A</th><th>B</th></tr><tr><td>1</td><td>2</td></tr></table>",
            img_path=None,
            bbox=None,
        )
    ]
    xlsx = tmp_path / "demo.xlsx"
    issues = write_excel(tables, xlsx)
    assert xlsx.is_file()
    assert issues == []
    wb = load_workbook(xlsx)
    assert len(wb.sheetnames) == 1
    ws = wb.active
    # data starts at row 4 when no footnote
    assert ws.cell(4, 1).value == "A"
    assert ws.cell(5, 2).value == "2"
